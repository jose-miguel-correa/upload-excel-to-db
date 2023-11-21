"""
import os

xls_path = './files'
print("Ruta a procesar: ", xls_path)

files = os.listdir(xls_path)
print("Archivos en carpeta: ", files)

found_files = [file for file in files if 'bd' in file.lower() or 'vpeo' in file.lower()]
print("Archivos a procesar: ", found_files)

for files in found_files:
    if 'bf' in files.lower():
        print(1)
    else:
        print("No se puede asignar código")
"""

import pandas as pd
import openpyxl
import xlrd
from connection import conn, cursor

""""""
# Remover autofiltros
#xlrd.open_workbook('./files/BD_STAFF_SEM_41.xlsx')
wb = openpyxl.load_workbook('./files/BD_STAFF_SEM_41.xlsx', read_only=True)
wb.save("./files/yuyumi.xlsx")


#print(ws['A5'])
#workbook = openpyxl.Workbook('./files/BD_STAFF_SEM_41.xlsx', write_only=False)
#workbook.copy_worksheet(from_worksheet='Staff  sin Cttos final <22')

"""
# Assuming 'Sheet1' is the name of your sheet
sheet = workbook['Staff  sin Cttos final <22']
print("uno")
# Remove filters from all columns
for sheet in workbook.sheetnames:
    ws = workbook[sheet]
    ws.auto_filter.ref = None
print("dos")
# Guarda archivo sin filtros
workbook.save('BD STAFF SEM 41.xls')
print("tres")

xls_path = './files/BD STAFF SEM 41.xls'
sheet = "Staff  sin Cttos final <22"
table_name = 'BD_STAFF_SEM_41'
print("cuatro")
df = pd.read_excel(xls_path, dtype=str, engine="openpyxl", sheet_name=sheet, skiprows=1)
print(df)
"""






